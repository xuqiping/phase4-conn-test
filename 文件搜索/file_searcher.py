import tkinter as tk
from tkinter import ttk, filedialog, scrolledtext, messagebox
import os
import threading
import queue
from concurrent.futures import ThreadPoolExecutor, as_completed
import time

try:
    import openpyxl
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False


# --- 现代深色配色方案 ---
THEME = {
    "bg_app": "#111827",       # 应用主背景 (深紫灰)
    "bg_card": "#1F2937",      # 卡片背景
    "bg_input": "#374151",     # 输入框背景
    "text_main": "#F9FAFB",    # 主文字颜色 (白)
    "text_sub": "#9CA3AF",     # 辅助文字颜色 (灰)
    "primary": "#3B82F6",      # 主色调/导出按钮 (蓝)
    "primary_hover": "#2563EB",
    "success": "#10B981",      # 开始按钮 (绿)
    "success_hover": "#059669",
    "danger": "#EF4444",       # 停止按钮 (红)
    "danger_hover": "#DC2626",
    "highlight": "#F59E0B",    # 高亮颜色 (黄/橙)
    "border": "#4B5563"
}


class RoundedButton(tk.Canvas):
    """自定义抗锯齿圆角按钮"""
    def __init__(self, parent, text, width, height, radius, color, hover_color, command=None, font=("Microsoft YaHei", 10, "bold")):
        super().__init__(parent, width=width, height=height, bg=parent["bg"], highlightthickness=0)
        self.command = command
        self.color = color
        self.hover_color = hover_color
        self.radius = radius
        self.text = text
        self.font = font
        self._hovered = False
        self._pressed = False

        self.rect_id = self._draw_round_rect(0, 0, width, height, radius, color)
        self.text_id = self.create_text(width/2, height/2, text=text, fill=THEME["text_main"], font=font)

        self.bind("<Enter>", self.on_enter)
        self.bind("<Leave>", self.on_leave)
        self.bind("<ButtonPress-1>", self.on_press)
        self.bind("<ButtonRelease-1>", self.on_release)

    def _draw_round_rect(self, x, y, w, h, r, color):
        self.create_arc(x, y, x+2*r, y+2*r, start=90, extent=90, fill=color, outline=color, tags="bg")
        self.create_arc(x+w-2*r, y, x+w, y+2*r, start=0, extent=90, fill=color, outline=color, tags="bg")
        self.create_arc(x, y+h-2*r, x+2*r, y+h, start=180, extent=90, fill=color, outline=color, tags="bg")
        self.create_arc(x+w-2*r, y+h-2*r, x+w, y+h, start=270, extent=90, fill=color, outline=color, tags="bg")
        self.create_rectangle(x+r, y, x+w-r, y+h, fill=color, outline=color, tags="bg")
        self.create_rectangle(x, y+r, x+w, y+h-r, fill=color, outline=color, tags="bg")
        return "bg"

    def on_enter(self, event):
        self._hovered = True
        self.itemconfig(self.rect_id, fill=self.hover_color, outline=self.hover_color)
        self.config(cursor="hand2")

    def on_leave(self, event):
        self._hovered = False
        self._pressed = False
        self.itemconfig(self.rect_id, fill=self.color, outline=self.color)
        self.config(cursor="")

    def on_press(self, event):
        self._pressed = True
        self.move(self.text_id, 0, 2)

    def on_release(self, event):
        self._pressed = False
        self.move(self.text_id, 0, -2)
        if self.command:
            self.command()

    def config_state(self, state):
        if state == tk.DISABLED:
            self.itemconfig(self.rect_id, fill=THEME["border"], outline=THEME["border"])
            self.itemconfig(self.text_id, fill=THEME["text_sub"])
            self.unbind("<Enter>")
            self.unbind("<Leave>")
            self.unbind("<ButtonPress-1>")
            self.unbind("<ButtonRelease-1>")
        else:
            self.itemconfig(self.rect_id, fill=self.color, outline=self.color)
            self.itemconfig(self.text_id, fill=THEME["text_main"])
            self.bind("<Enter>", self.on_enter)
            self.bind("<Leave>", self.on_leave)
            self.bind("<ButtonPress-1>", self.on_press)
            self.bind("<ButtonRelease-1>", self.on_release)


class FlatEntry(tk.Frame):
    """扁平化输入框"""
    def __init__(self, parent, textvariable, placeholder="", width=40):
        super().__init__(parent, bg=THEME["bg_input"], padx=10, pady=8)
        self.placeholder = placeholder
        self.textvariable = textvariable
        self.has_placeholder = True

        self.entry = tk.Entry(self, textvariable=textvariable, bg=THEME["bg_input"],
                              fg=THEME["text_sub"], insertbackground=THEME["text_main"],
                              relief="flat", font=("Microsoft YaHei", 10), borderwidth=0,
                              highlightthickness=0)
        self.entry.pack(fill=tk.X, expand=True)

        if placeholder and not textvariable.get():
            self.entry.insert(0, placeholder)
            self.has_placeholder = True

        self.entry.bind("<FocusIn>", self._on_focus_in)
        self.entry.bind("<FocusOut>", self._on_focus_out)
        self.entry.bind("<KeyRelease>", self._on_key_release)

    def _on_focus_in(self, event):
        if self.has_placeholder:
            self.entry.delete(0, tk.END)
            self.entry.config(fg=THEME["text_main"])
            self.has_placeholder = False

    def _on_focus_out(self, event):
        if not self.entry.get():
            self.entry.insert(0, self.placeholder)
            self.entry.config(fg=THEME["text_sub"])
            self.has_placeholder = True

    def _on_key_release(self, event):
        if self.entry.get() and self.entry.get() != self.placeholder:
            self.textvariable.set(self.entry.get())

    def get(self):
        if self.has_placeholder:
            return ""
        return self.entry.get()

    def clear(self):
        self.entry.delete(0, tk.END)
        self.entry.insert(0, self.placeholder)
        self.entry.config(fg=THEME["text_sub"])
        self.has_placeholder = True


class ModernFileSearcher:
    def __init__(self, root):
        self.root = root
        self.root.title("文件搜索工具 Pro")
        self.root.geometry("1000x750")
        self.root.minsize(800, 600)
        self.root.configure(bg=THEME["bg_app"])

        self.searching = False
        self.result_queue = queue.Queue()
        self.file_list = []
        self.processed_count = 0
        self.found_count = 0

        self.setup_ui()
        self.setup_tags()

    def setup_ui(self):
        main_container = tk.Frame(self.root, bg=THEME["bg_app"])
        main_container.pack(fill=tk.BOTH, expand=True, padx=24, pady=24)

        # 顶部标题区
        header_frame = tk.Frame(main_container, bg=THEME["bg_app"])
        header_frame.pack(fill=tk.X, pady=(0, 20))
        ttk.Label(header_frame, text="🔍 文件搜索工具 Pro",
                  font=("Microsoft YaHei", 18, "bold"), foreground=THEME["primary"],
                  background=THEME["bg_app"]).pack(side=tk.LEFT)

        # 搜索配置卡片
        config_card = tk.Frame(main_container, bg=THEME["bg_card"],
                                highlightthickness=1, highlightbackground=THEME["border"],
                                highlightcolor=THEME["border"])
        config_card.pack(fill=tk.X, pady=(0, 20))

        config_inner = tk.Frame(config_card, bg=THEME["bg_card"], padx=20, pady=20)
        config_inner.pack(fill=tk.BOTH, expand=True)
        config_inner.columnconfigure(1, weight=1)
        config_inner.columnconfigure(3, weight=1)

        # 搜索目录
        ttk.Label(config_inner, text="📂 搜索目录:", font=("Microsoft YaHei", 10),
                  foreground=THEME["text_main"], background=THEME["bg_card"]).grid(row=0, column=0, sticky="w", pady=(0, 15), padx=(0, 10))

        self.dir_path = tk.StringVar()
        self.dir_entry = FlatEntry(config_inner, self.dir_path, "请选择或输入目标文件夹路径...")
        self.dir_entry.grid(row=0, column=1, sticky="ew", pady=(0, 15))

        browse_btn = RoundedButton(config_inner, "浏览", 80, 36, 6, THEME["primary"], THEME["primary_hover"], self.browse_directory)
        browse_btn.grid(row=0, column=2, padx=(10, 0), pady=(0, 15))

        # 搜索内容
        ttk.Label(config_inner, text="📝 搜索内容:", font=("Microsoft YaHei", 10),
                  foreground=THEME["text_main"], background=THEME["bg_card"]).grid(row=1, column=0, sticky="w", pady=(0, 15), padx=(0, 10))

        self.search_content = tk.StringVar()
        self.keyword_entry = FlatEntry(config_inner, self.search_content, "输入要搜索的关键词...")
        self.keyword_entry.grid(row=1, column=1, columnspan=2, sticky="ew", pady=(0, 15))

        # 搜索模式和文件类型
        options_frame = tk.Frame(config_inner, bg=THEME["bg_card"])
        options_frame.grid(row=2, column=0, columnspan=3, sticky="ew", pady=(0, 5))

        ttk.Label(options_frame, text="📊 搜索模式:", font=("Microsoft YaHei", 10),
                  foreground=THEME["text_main"], background=THEME["bg_card"]).pack(side=tk.LEFT, padx=(0, 10))

        self.search_mode = tk.StringVar(value="content")
        mode_cb = ttk.Combobox(options_frame, textvariable=self.search_mode,
                               values=["按文件内容", "按文件名/目录名", "混合模式"],
                               state="readonly", width=15, font=("Microsoft YaHei", 10))
        mode_cb.pack(side=tk.LEFT, padx=(0, 30))
        mode_cb.bind("<<ComboboxSelected>>", self._on_mode_changed)

        ttk.Label(options_frame, text="📄 文件类型:", font=("Microsoft YaHei", 10),
                  foreground=THEME["text_main"], background=THEME["bg_card"]).pack(side=tk.LEFT, padx=(0, 10))

        self.file_extensions = tk.StringVar(value="txt,py,md,js,json,html,css,xlsx")
        self.types_entry = FlatEntry(options_frame, self.file_extensions, "例如: txt, py, md")
        self.types_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))

        # 快捷类型按钮
        quick_frame = tk.Frame(options_frame, bg=THEME["bg_card"])
        quick_frame.pack(side=tk.LEFT)

        quick_types = [("全部", "*"), ("文本", "txt,py,md,js,json,html,css"), ("Office", "xlsx"), ("代码", "py,js,java,cpp")]
        for label, ext in quick_types:
            btn = tk.Label(quick_frame, text=label, bg=THEME["bg_input"], fg=THEME["text_main"],
                          font=("Microsoft YaHei", 9), padx=8, pady=4, cursor="hand2")
            btn.pack(side=tk.LEFT, padx=2)
            btn.bind("<Button-1>", lambda e, ext=ext: self.file_extensions.set(ext))

        # 操作按钮区
        btn_frame = tk.Frame(main_container, bg=THEME["bg_app"])
        btn_frame.pack(fill=tk.X, pady=(0, 15))

        self.search_btn = RoundedButton(btn_frame, "▶ 开始搜索", 120, 40, 8, THEME["success"], THEME["success_hover"], self.start_search)
        self.search_btn.pack(side=tk.LEFT, padx=(0, 10))

        self.stop_btn = RoundedButton(btn_frame, "⏹ 停止", 100, 40, 8, THEME["danger"], THEME["danger_hover"], self.stop_search)
        self.stop_btn.config_state(tk.DISABLED)

        self.export_btn = RoundedButton(btn_frame, "💾 导出结果", 120, 40, 8, THEME["primary"], THEME["primary_hover"], self.export_results)
        self.export_btn.config_state(tk.DISABLED)
        self.export_btn.pack(side=tk.RIGHT)

        # 进度条区
        progress_frame = tk.Frame(main_container, bg=THEME["bg_app"])
        progress_frame.pack(fill=tk.X, pady=(0, 15))

        self.progress = ttk.Progressbar(progress_frame, mode='determinate', length=200)
        self.progress.configure(style="Modern.Horizontal.TProgressbar")
        self.progress.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=4)

        self.progress_label = tk.Label(progress_frame, text="0%", bg=THEME["bg_app"],
                                       fg=THEME["text_main"], font=("Consolas", 10))
        self.progress_label.pack(side=tk.LEFT, padx=(15, 0))

        # 结果显示区
        result_card = tk.Frame(main_container, bg=THEME["bg_card"],
                               highlightthickness=1, highlightbackground=THEME["border"])
        result_card.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

        self.result_text = scrolledtext.ScrolledText(result_card, wrap=tk.WORD,
                                                     bg=THEME["bg_card"], fg=THEME["text_main"],
                                                     font=("Consolas", 10), relief="flat", borderwidth=0,
                                                     insertbackground=THEME["text_main"],
                                                     selectbackground=THEME["primary"],
                                                     selectforeground="#FFF",
                                                     padx=12, pady=12)
        self.result_text.pack(fill=tk.BOTH, expand=True)

        # 底部状态栏
        status_frame = tk.Frame(self.root, bg="#0F141E", padx=20, pady=8)
        status_frame.pack(side=tk.BOTTOM, fill=tk.X)

        self.status_var = tk.StringVar(value="就绪 | 等待操作")
        tk.Label(status_frame, textvariable=self.status_var, bg="#0F141E",
                fg=THEME["text_sub"], font=("Microsoft YaHei", 9), anchor="w").pack(side=tk.LEFT)

        self.stats_var = tk.StringVar(value="找到: 0 | 已扫描: 0")
        tk.Label(status_frame, textvariable=self.stats_var, bg="#0F141E",
                fg=THEME["text_sub"], font=("Microsoft YaHei", 9)).pack(side=tk.RIGHT)

        # 配置进度条样式
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Modern.Horizontal.TProgressbar", troughcolor=THEME["bg_input"],
                       background=THEME["primary"], thickness=8)

    def setup_tags(self):
        self.result_text.tag_config("icon", foreground=THEME["primary"], font=("Microsoft YaHei", 10))
        self.result_text.tag_config("path", foreground=THEME["text_main"], font=("Microsoft YaHei", 10, "bold"))
        self.result_text.tag_config("label", foreground=THEME["highlight"], font=("Microsoft YaHei", 10))
        self.result_text.tag_config("locations", foreground=THEME["success"], font=("Microsoft YaHei", 10))
        self.result_text.tag_config("more", foreground=THEME["text_sub"], font=("Microsoft YaHei", 9))
        self.result_text.tag_config("separator", foreground=THEME["border"])

    def _on_mode_changed(self, event=None):
        mode_text = self.search_mode.get()
        if mode_text == "按文件名/目录名":
            self.search_mode.set("name")
        elif mode_text == "按文件内容":
            self.search_mode.set("content")
        elif mode_text == "混合模式":
            self.search_mode.set("mixed")

    def browse_directory(self):
        directory = filedialog.askdirectory()
        if directory:
            self.dir_path.set(directory)

    def collect_files(self, search_dir, ext_list):
        files = []
        for root, dirs, filenames in os.walk(search_dir):
            if not self.searching:
                break
            for filename in filenames:
                if ext_list:
                    file_ext = os.path.splitext(filename)[1].lower().lstrip('.')
                    if file_ext not in ext_list:
                        continue
                files.append(os.path.join(root, filename))
        return files

    def collect_all_paths(self, search_dir):
        paths = []
        for root, dirs, filenames in os.walk(search_dir):
            if not self.searching:
                break
            for dir_name in dirs:
                paths.append(('dir', os.path.join(root, dir_name)))
            for filename in filenames:
                paths.append(('file', os.path.join(root, filename)))
        return paths

    def search_text_file(self, file_path, search_content):
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
                if search_content in content:
                    lines = content.split('\n')
                    line_numbers = []
                    for i, line in enumerate(lines, 1):
                        if search_content in line:
                            line_numbers.append(str(i))
                    return file_path, line_numbers, "text"
        except Exception:
            pass
        return None

    def search_xlsx_file(self, file_path, search_content):
        if not HAS_XLSX:
            return None
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            matches = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                    if not self.searching:
                        break
                    for col_idx, cell_value in enumerate(row, 1):
                        if cell_value and search_content in str(cell_value):
                            col_letter = openpyxl.utils.get_column_letter(col_idx)
                            matches.append(f"{sheet_name}!{col_letter}{row_idx}")
            wb.close()
            if matches:
                return file_path, matches, "xlsx"
        except Exception:
            pass
        return None

    def search_single_file(self, file_path, search_content):
        if not self.searching:
            return None

        ext = os.path.splitext(file_path)[1].lower()

        if ext == '.xlsx':
            return self.search_xlsx_file(file_path, search_content)
        else:
            return self.search_text_file(file_path, search_content)

    def process_results(self):
        batch = []
        last_update = time.time()

        while self.searching or not self.result_queue.empty():
            try:
                result = self.result_queue.get(timeout=0.1)
                if result:
                    batch.append(result)
                    self.found_count += 1
            except queue.Empty:
                pass

            current_time = time.time()
            if batch and (current_time - last_update > 0.1 or len(batch) >= 10):
                self.root.after(0, self.append_results_batch, batch.copy())
                batch = []
                last_update = current_time

        if batch:
            self.root.after(0, self.append_results_batch, batch)

    def append_results_batch(self, results):
        for result in results:
            result_type = result[0]

            if result_type == 'name_match':
                path_type, full_path = result[1], result[2]
                if path_type == 'dir':
                    self.result_text.insert(tk.END, "📁 ", "icon")
                    self.result_text.insert(tk.END, f"{full_path}\n", "path")
                    self.result_text.insert(tk.END, "   📂 目录名匹配\n", "label")
                else:
                    self.result_text.insert(tk.END, "📄 ", "icon")
                    self.result_text.insert(tk.END, f"{full_path}\n", "path")
                    self.result_text.insert(tk.END, "   🏷️  文件名匹配\n", "label")
            else:
                file_path, locations, file_type = result
                self.result_text.insert(tk.END, "📄 ", "icon")
                self.result_text.insert(tk.END, f"{file_path}\n", "path")

                if file_type == "xlsx":
                    self.result_text.insert(tk.END, "   📊 单元格: ", "label")
                    self.result_text.insert(tk.END, f"{', '.join(locations[:10])}", "locations")
                    if len(locations) > 10:
                        self.result_text.insert(tk.END, f" ... (共{len(locations)}处)", "more")
                else:
                    self.result_text.insert(tk.END, "   📝 行号: ", "label")
                    self.result_text.insert(tk.END, f"{', '.join(locations[:10])}", "locations")
                    if len(locations) > 10:
                        self.result_text.insert(tk.END, f" ... (共{len(locations)}处)", "more")

            self.result_text.insert(tk.END, "─" * 80 + "\n", "separator")

        self.result_text.see(tk.END)
        self.stats_var.set(f"找到: {self.found_count} | 已扫描: {self.processed_count}/{len(self.file_list)}")

    def start_search(self):
        if not self.dir_path.get():
            self.status_var.set("请先选择搜索目录")
            messagebox.showwarning("提示", "请先选择搜索目录")
            return
        if not os.path.isdir(self.dir_path.get()):
            self.status_var.set("目录不存在")
            messagebox.showwarning("提示", "目录不存在")
            return
        if not self.search_content.get():
            self.status_var.set("请输入搜索内容")
            messagebox.showwarning("提示", "请输入搜索内容")
            return

        self.searching = True
        self.search_btn.config_state(tk.DISABLED)
        self.stop_btn.config_state(tk.NORMAL)
        self.export_btn.config_state(tk.DISABLED)
        self.result_text.delete(1.0, tk.END)
        self.processed_count = 0
        self.found_count = 0
        self.progress['value'] = 0
        self.progress_label['text'] = "0%"

        mode = self.search_mode.get()
        if mode not in ["content", "name", "mixed"]:
            self._on_mode_changed()

        extensions = self.file_extensions.get().strip()
        if extensions == "*" or not extensions or "输入" in extensions:
            ext_list = None
        else:
            ext_list = [ext.strip().lower() for ext in extensions.split(",")]

        self.status_var.set("正在收集文件列表...")
        self.root.update()

        if self.search_mode.get() == 'name':
            self.file_list = self.collect_all_paths(self.dir_path.get())
        else:
            self.file_list = self.collect_files(self.dir_path.get(), ext_list)

        if not self.file_list:
            self.status_var.set("未找到符合条件的文件")
            self.search_finished()
            return

        self.stats_var.set(f"找到: 0 | 已扫描: 0/{len(self.file_list)}")

        self.search_thread = threading.Thread(target=self.run_search, args=(self.search_mode.get(),))
        self.search_thread.daemon = True
        self.search_thread.start()

        self.result_processor_thread = threading.Thread(target=self.process_results)
        self.result_processor_thread.daemon = True
        self.result_processor_thread.start()

    def run_search(self, mode):
        search_content = self.search_content.get()
        total_items = len(self.file_list)

        if mode == 'name':
            self.status_var.set(f"正在搜索 {total_items} 个路径...")
            for item in self.file_list:
                if not self.searching:
                    break
                self.processed_count += 1
                path_type, full_path = item
                name = os.path.basename(full_path)
                if search_content in name:
                    self.result_queue.put(('name_match', path_type, full_path))
                progress = (self.processed_count / total_items) * 100
                self.root.after(0, self.update_progress, progress)
        elif mode == 'content':
            self.status_var.set(f"正在搜索 {total_items} 个文件...")
            max_workers = min(16, (os.cpu_count() or 4) * 2)

            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                futures = {executor.submit(self.search_single_file, file_path, search_content): file_path
                          for file_path in self.file_list}

                for future in as_completed(futures):
                    if not self.searching:
                        for f in futures:
                            f.cancel()
                        break

                    self.processed_count += 1
                    progress = (self.processed_count / total_items) * 100
                    self.root.after(0, self.update_progress, progress)

                    try:
                        result = future.result()
                        if result:
                            self.result_queue.put(result)
                    except Exception:
                        pass
        else:
            self.status_var.set(f"正在混合搜索 {total_items} 个文件...")
            search_dir = self.dir_path.get()
            all_paths = self.collect_all_paths(search_dir)

            name_matches = []
            for item in all_paths:
                if not self.searching:
                    break
                path_type, full_path = item
                name = os.path.basename(full_path)
                if search_content in name:
                    name_matches.append(('name_match', path_type, full_path))

            for match in name_matches:
                self.result_queue.put(match)
                self.found_count += 1

            max_workers = min(16, (os.cpu_count() or 4) * 2)
            total_for_content = len(self.file_list)

            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                futures = {executor.submit(self.search_single_file, file_path, search_content): file_path
                          for file_path in self.file_list}

                for future in as_completed(futures):
                    if not self.searching:
                        for f in futures:
                            f.cancel()
                        break

                    self.processed_count += 1
                    total_so_far = len(name_matches) + total_for_content
                    progress = (self.processed_count / total_for_content) * 100 if total_for_content > 0 else 100
                    self.root.after(0, self.update_progress, progress)

                    try:
                        result = future.result()
                        if result:
                            self.result_queue.put(result)
                    except Exception:
                        pass

            self.file_list = name_matches + self.file_list

        if self.searching:
            self.status_var.set(f"搜索完成! 找到 {self.found_count} 个匹配")
        else:
            self.status_var.set(f"搜索已停止! 找到 {self.found_count} 个匹配")

        self.root.after(0, self.search_finished)

    def update_progress(self, value):
        self.progress['value'] = value
        self.progress_label['text'] = f"{value:.1f}%"
        self.stats_var.set(f"找到: {self.found_count} | 已扫描: {self.processed_count}/{len(self.file_list)}")

    def stop_search(self):
        self.searching = False
        self.status_var.set("正在停止...")

    def search_finished(self):
        self.searching = False
        self.search_btn.config_state(tk.NORMAL)
        self.stop_btn.config_state(tk.DISABLED)
        if self.found_count > 0:
            self.export_btn.config_state(tk.NORMAL)
        self.stats_var.set(f"找到: {self.found_count} | 已扫描: {self.processed_count}/{len(self.file_list)}")

    def export_results(self):
        file_path = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")]
        )
        if file_path:
            try:
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(self.result_text.get(1.0, tk.END))
                self.status_var.set(f"结果已导出到: {file_path}")
                messagebox.showinfo("成功", f"结果已导出到:\n{file_path}")
            except Exception as e:
                self.status_var.set(f"导出失败: {str(e)}")
                messagebox.showerror("错误", f"导出失败:\n{str(e)}")


def main():
    root = tk.Tk()
    app = ModernFileSearcher(root)
    root.mainloop()


if __name__ == "__main__":
    main()
